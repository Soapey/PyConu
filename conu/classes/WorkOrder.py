from datetime import datetime, date
from conu.classes.PriorityLevel import PriorityLevel
from conu.classes.WorkOrderItem import WorkOrderItem
from conu.classes.WorkOrderAssignee import WorkOrderAssignee
from conu.classes.Item import Item
from conu.classes.Assignee import Assignee
from conu.classes.Site import Site
from conu.classes.Department import Department
from conu.classes.User import User
from conu.db.SQLiteConnection import SQLiteConnection
from conu.db.helpers import select_by_attrs_dict, format_nullable_database_date
from conu.helpers import select_directory, hex_to_rgb
from conu.ui.components.Notification import SuccessNotification, ErrorNotification
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, Color
from openpyxl.utils import range_boundaries
import win32com.client as win32
import os


class WorkOrder:
    def __init__(
        self,
        id: int,
        site_id: int,
        department_id: int,
        prioritylevel_id: int,
        task_description: str,
        comments: str,
        date_created: date,
        date_allocated: date,
        raisedby_user_id: int,
        date_completed: date,
        purchase_order_number: str,
        close_out_comments: str,
    ) -> None:
        self.id = id
        self.site_id = site_id
        self.department_id = department_id
        self.prioritylevel_id = prioritylevel_id
        self.task_description = task_description
        self.comments = comments
        self.date_created = date_created
        self.date_allocated = date_allocated
        self.raisedby_user_id = raisedby_user_id
        self.date_completed = date_completed
        self.purchase_order_number = purchase_order_number
        self.close_out_comments = close_out_comments

    def __repr__(self):
        attrs = ", ".join(f"{k}={v!r}" for k, v in self.__dict__.items())
        return f"{self.__class__.__name__}({attrs})"

    def __str__(self) -> str:
        return self.__repr__()

    def due_listingview_items(self):

        items = select_by_attrs_dict(Item)
        workorderitems = select_by_attrs_dict(WorkOrderItem)

        relevant_workorderitems = {
            workorderitem.id: workorderitem
            for workorderitem in workorderitems.values()
            if workorderitem.workorder_id == self.id
        }

        item_list = list()
        for workorderitem in relevant_workorderitems.values():
            item = items[workorderitem.item_id]
            item_list.append(item.name)

        return ", ".join(item_list)

    def due_listingview_assignees(self):

        assignees = select_by_attrs_dict(Assignee)
        workorderassignees = select_by_attrs_dict(WorkOrderAssignee)

        relevant_workorderassignees = {
            workorderassignee.id: workorderassignee
            for workorderassignee in workorderassignees.values()
            if workorderassignee.workorder_id == self.id
        }

        assignee_list = list()
        for workorderassignee in relevant_workorderassignees.values():
            assignee = assignees[workorderassignee.assignee_id]
            assignee_list.append(assignee.name)

        return ", ".join(assignee_list)

    def due_listingview_summary(self):
        return self.task_description

    def is_due(self) -> bool:

        prioritylevels = select_by_attrs_dict(PriorityLevel)
        priority_level = prioritylevels[self.prioritylevel_id]

        days_until_overdue = priority_level.days_until_overdue

        return (
            not self.date_completed
            and (date.today() - self.date_allocated).days >= days_until_overdue
        )

    @classmethod
    def convert_rows_to_instances(cls, rows):

        return {
            row[0]: cls(
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
                row[5],
                format_nullable_database_date(row[6]),
                format_nullable_database_date(row[7]),
                row[8],
                format_nullable_database_date(row[9]),
                row[10],
                row[11],
            )
            for row in rows
        }

    @classmethod
    def get_by_user_departments(cls, user_id):

        with SQLiteConnection() as cur:

            rows = cur.execute(
                """
                SELECT 
                    * 
                FROM 
                    workorder 
                WHERE 
                    workorder.department_id IN 
                        (SELECT userdepartment.department_id FROM userdepartment WHERE userdepartment.user_id = ?);""",
                (user_id,),
            ).fetchall()

        return cls.convert_rows_to_instances(rows)

    @classmethod
    def get_listingview_table_data(cls, main_window):

        current_user = main_window.current_user

        if not current_user:
            return

        with SQLiteConnection() as cur:
            rows = cur.execute(
                """
            SELECT
                workorder.id,
                site.name,
                department.name,
                prioritylevel.name,
                workorder.task_description,
                workorder.comments,
                GROUP_CONCAT(DISTINCT item.name),
                GROUP_CONCAT(DISTINCT assignee.name),
                strftime('%d-%m-%Y', workorder.date_allocated),
                strftime('%d-%m-%Y', workorder.date_completed),
                workorder.close_out_comments,
                user.first_name || ' ' || user.last_name,
                strftime('%d-%m-%Y', workorder.date_created)
            FROM
                workorder
                JOIN site ON workorder.site_id = site.id
                JOIN department ON workorder.department_id = department.id
                JOIN prioritylevel ON workorder.prioritylevel_id = prioritylevel.id
                JOIN user ON workorder.raisedby_user_id = user.id
                LEFT JOIN workorderitem ON workorder.id = workorderitem.workorder_id
                LEFT JOIN item ON workorderitem.item_id = item.id
                LEFT JOIN workorderassignee ON workorder.id = workorderassignee.workorder_id
                LEFT JOIN assignee ON workorderassignee.assignee_id = assignee.id
            WHERE
                workorder.department_id IN (
                    SELECT department_id
                    FROM userdepartment
                    WHERE user_id = ?
                )
            GROUP BY
                workorder.id
            ORDER BY
                workorder.id ASC;""",
                (current_user.id,),
            ).fetchall()

        return rows

    @classmethod
    def get(cls):

        with SQLiteConnection() as cur:

            rows = cur.execute("SELECT * FROM workorder;").fetchall()

        return {
            row[0]: cls(
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
                row[5],
                format_nullable_database_date(row[6]),
                format_nullable_database_date(row[7]),
                row[8],
                format_nullable_database_date(row[9]),
                row[10],
                row[11],
            )
            for row in rows
        }

    def site(self):

        sites = select_by_attrs_dict(Site, {"id": self.site_id})

        if sites:
            return sites[self.site_id]

        return None

    def department(self):

        departments = select_by_attrs_dict(Department, {"id": self.department_id})

        if departments:
            return departments[self.department_id]

        return None

    def prioritylevel(self):

        prioritylevels = select_by_attrs_dict(
            PriorityLevel, {"id": self.prioritylevel_id}
        )

        if prioritylevels:
            return prioritylevels[self.prioritylevel_id]

        return None

    def items(self):

        with SQLiteConnection() as cur:

            rows = cur.execute(
                """
            SELECT 
                * 
            FROM 
                item 
            WHERE item.id IN (
                SELECT 
                    workorderitem.item_id 
                FROM 
                    workorderitem 
                WHERE 
                    workorderitem.workorder_id = ?
                );
            """,
                (self.id,),
            ).fetchall()

        return Item.convert_rows_to_instances(rows)

    def assignees(self):

        with SQLiteConnection() as cur:

            rows = cur.execute(
                """
            SELECT 
                * 
            FROM 
                assignee 
            WHERE assignee.id IN (
                SELECT 
                    workorderassignee.assignee_id 
                FROM 
                    workorderassignee
                WHERE 
                    workorderassignee.workorder_id = ?
                );
            """,
                (self.id,),
            ).fetchall()

        return Assignee.convert_rows_to_instances(rows)

    def raisedby(self):

        users = select_by_attrs_dict(User, {"id": self.raisedby_user_id})

        if users:
            return users[self.raisedby_user_id]

        return None

    def save(self, print_on_save: bool = False):

        COLUMNS = 8
        DARK_BLUE = "#1D3557"
        LIGHT_BLUE = "#457B9D"

        DARK_BLUE_RGB = f"FF{DARK_BLUE.lstrip('#')}"
        LIGHT_BLUE_RGB = f"FF{LIGHT_BLUE.lstrip('#')}"

        SMALL_HEADER_BACKGROUND = PatternFill(
            start_color=LIGHT_BLUE_RGB,
            end_color=LIGHT_BLUE_RGB,
            fill_type="solid",
        )
        SMALL_HEADER_FONT = Font(name="Helvetica", size=10, color="FFFFFF", bold=True)
        VALUE_FONT = Font(name="Helvetica", size=14, color="000000")

        SUB_HEADER_BACKGROUND = PatternFill(
            start_color=DARK_BLUE_RGB,
            end_color=DARK_BLUE_RGB,
            fill_type="solid",
        )
        SUB_HEADER_FONT = Font(
            name="Helvetica", size=16, color="FFFFFF", italic=True, bold=True
        )

        workbook = openpyxl.Workbook()

        sheet = workbook.worksheets[0]

        column_range = sheet.iter_cols(min_col=1, max_col=8)
        for column in column_range:
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = 15.71

        border_style = Side(style="thin", color="000000")
        border = Border(
            top=border_style, left=border_style, right=border_style, bottom=border_style
        )

        ###################################
        row = 1
        title_label = sheet.cell(row, 1)
        title_label.value = "WORK ORDER FORM"
        title_label.font = Font(
            name="Helvetica", size=24, italic=True, bold=True, color=DARK_BLUE
        )
        ###################################
        row += 3
        number_label = sheet.cell(row, 1)
        number_label.value = "Number"
        number_label.fill = SMALL_HEADER_BACKGROUND
        number_label.font = SMALL_HEADER_FONT
        number_label.alignment = Alignment(horizontal="right")
        number_label.border = border
        sheet.merge_cells(f"A{row}:B{row}")
        number_value = sheet.cell(row, 3)
        number_value.value = self.id
        number_value.alignment = Alignment(horizontal="left")
        number_value.font = VALUE_FONT
        number_value.border = border
        sheet.merge_cells(f"C{row}:H{row}")

        row += 1
        date_created_label = sheet.cell(row, 1)
        date_created_label.value = "Date Created"
        date_created_label.fill = SMALL_HEADER_BACKGROUND
        date_created_label.font = SMALL_HEADER_FONT
        date_created_label.alignment = Alignment(horizontal="right")
        date_created_label.border = border
        sheet.merge_cells(f"A{row}:B{row}")
        date_created_value = sheet.cell(row, 3)
        date_created_value.value = datetime.strftime(self.date_created, "%d-%m-%Y")
        date_created_value.font = VALUE_FONT
        date_created_value.border = border
        sheet.merge_cells(f"C{row}:H{row}")

        row += 1
        raisedby = self.raisedby()
        raisedby_label = sheet.cell(row, 1)
        raisedby_label.value = "Raised By"
        raisedby_label.fill = SMALL_HEADER_BACKGROUND
        raisedby_label.font = SMALL_HEADER_FONT
        raisedby_label.alignment = Alignment(horizontal="right")
        raisedby_label.border = border
        sheet.merge_cells(f"A{row}:B{row}")
        raisedby_value = sheet.cell(row, 3)
        raisedby_value.value = f"{raisedby.first_name} {raisedby.last_name}"
        raisedby_value.font = VALUE_FONT
        raisedby_value.border = border
        sheet.merge_cells(f"C{row}:D{row}")
        contactemail_label = sheet.cell(row, 5)
        contactemail_label.value = "Contact Email"
        contactemail_label.fill = SMALL_HEADER_BACKGROUND
        contactemail_label.font = SMALL_HEADER_FONT
        contactemail_label.alignment = Alignment(horizontal="right")
        contactemail_label.border = border
        sheet.merge_cells(f"E{row}:F{row}")
        contactemail_value = sheet.cell(row, 7)
        contactemail_value.value = f"{raisedby.email_address}"
        contactemail_value.font = VALUE_FONT
        contactemail_value.border = border
        sheet.merge_cells(f"G{row}:H{row}")

        row += 2
        site = self.site()
        site_label = sheet.cell(row, 1)
        site_label.value = "Site"
        site_label.fill = SMALL_HEADER_BACKGROUND
        site_label.font = SMALL_HEADER_FONT
        site_label.alignment = Alignment(horizontal="right")
        site_label.border = border
        sheet.merge_cells(f"A{row}:B{row}")
        site_value = sheet.cell(row, 3)
        site_value.value = site.name
        site_value.font = VALUE_FONT
        site_value.border = border
        sheet.merge_cells(f"C{row}:H{row}")

        row += 1
        department = self.department()
        department_label = sheet.cell(row, 1)
        department_label.value = "Department"
        department_label.fill = SMALL_HEADER_BACKGROUND
        department_label.font = SMALL_HEADER_FONT
        department_label.alignment = Alignment(horizontal="right")
        department_label.border = border
        sheet.merge_cells(f"A{row}:B{row}")
        department_value = sheet.cell(row, 3)
        department_value.value = department.name
        department_value.font = VALUE_FONT
        department_value.border = border
        sheet.merge_cells(f"C{row}:H{row}")

        row += 1
        prioritylevel = self.prioritylevel()
        prioritylevel_label = sheet.cell(row, 1)
        prioritylevel_label.value = "Priority Level"
        prioritylevel_label.fill = SMALL_HEADER_BACKGROUND
        prioritylevel_label.font = SMALL_HEADER_FONT
        prioritylevel_label.alignment = Alignment(horizontal="right")
        prioritylevel_label.border = border
        sheet.merge_cells(f"A{row}:B{row}")
        prioritylevel_value = sheet.cell(row, 3)
        prioritylevel_value.value = prioritylevel.name
        prioritylevel_value.font = VALUE_FONT
        prioritylevel_value.border = border
        sheet.merge_cells(f"C{row}:H{row}")

        row += 1
        purchaseordernumber_label = sheet.cell(row, 1)
        purchaseordernumber_label.value = "PO Number"
        purchaseordernumber_label.fill = SMALL_HEADER_BACKGROUND
        purchaseordernumber_label.font = SMALL_HEADER_FONT
        purchaseordernumber_label.alignment = Alignment(horizontal="right")
        purchaseordernumber_label.border = border
        sheet.merge_cells(f"A{row}:B{row}")
        purchaseordernumber_value = sheet.cell(row, 3)
        purchaseordernumber_value.value = self.purchase_order_number
        purchaseordernumber_value.font = VALUE_FONT
        purchaseordernumber_value.border = border
        sheet.merge_cells(f"C{row}:H{row}")

        row += 1
        dateallocated_label = sheet.cell(row, 1)
        dateallocated_label.value = "Date Allocated"
        dateallocated_label.fill = SMALL_HEADER_BACKGROUND
        dateallocated_label.font = SMALL_HEADER_FONT
        dateallocated_label.alignment = Alignment(horizontal="right")
        dateallocated_label.border = border
        sheet.merge_cells(f"A{row}:B{row}")
        dateallocated_value = sheet.cell(row, 3)
        dateallocated_value.value = datetime.strftime(self.date_allocated, "%d-%m-%Y")
        dateallocated_value.font = VALUE_FONT
        dateallocated_value.border = border
        sheet.merge_cells(f"C{row}:D{row}")
        overduefrom_label = sheet.cell(row, 5)
        overduefrom_label.value = "Overdue From"
        overduefrom_label.fill = SMALL_HEADER_BACKGROUND
        overduefrom_label.font = SMALL_HEADER_FONT
        overduefrom_label.alignment = Alignment(horizontal="right")
        overduefrom_label.border = border
        sheet.merge_cells(f"E{row}:F{row}")
        overduefrom_value = sheet.cell(row, 7)
        overduefrom_value.value = datetime.strftime(self.date_allocated, "%d-%m-%Y")
        overduefrom_value.font = VALUE_FONT
        overduefrom_value.border = border
        sheet.merge_cells(f"G{row}:H{row}")

        ###################################
        row += 2
        taskdescription_label = sheet.cell(row, 1)
        taskdescription_label.value = "TASK DESCRIPTION"
        taskdescription_label.border = border
        taskdescription_label.alignment = Alignment(horizontal="center")
        taskdescription_label.fill = SUB_HEADER_BACKGROUND
        taskdescription_label.font = SUB_HEADER_FONT
        sheet.merge_cells(f"A{row}:H{row}")
        row += 1
        taskdescription_value = sheet.cell(row, 1)
        taskdescription_value.value = self.task_description
        taskdescription_value.font = Font(name="Helvetica", size=12)
        taskdescription_value.alignment = Alignment(horizontal="left", wrap_text=True)
        sheet.merge_cells(f"A{row}:H{row}")

        lines = taskdescription_value.value.count("\n")
        row_height = (lines * 17) + 6
        sheet.row_dimensions[row].height = max(15.00, row_height)

        start_col, start_row, end_col, end_row = range_boundaries(f"A{row}:H{row}")
        for _row in sheet.iter_rows(
            min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col
        ):
            for cell in _row:
                cell.border = border

        ###################################
        row += 2
        items_label = sheet.cell(row, 1)
        items_label.value = "ITEMS"
        items_label.border = border
        items_label.alignment = Alignment(horizontal="center")
        items_label.fill = SUB_HEADER_BACKGROUND
        items_label.font = SUB_HEADER_FONT
        sheet.merge_cells(f"A{row}:H{row}")
        sections = 4
        row += 1
        col = 1
        for index, item in enumerate(self.items().values()):
            item_cell = sheet.cell(row, col)
            item_cell.value = item.name
            item_cell.font = Font(name="Helvetica", size=12)
            item_cell.border = border
            sheet.merge_cells(
                start_row=row,
                end_row=row,
                start_column=col,
                end_column=col + ((COLUMNS // sections) - 1),
            )
            col += 2
            if (index + 1) % sections == 0:
                col = 1
                row += 1
        ###################################
        row += 2
        assignees_label = sheet.cell(row, 1)
        assignees_label.value = "ASSIGNEES"
        assignees_label.border = border
        assignees_label.alignment = Alignment(horizontal="center")
        assignees_label.fill = SUB_HEADER_BACKGROUND
        assignees_label.font = SUB_HEADER_FONT
        sheet.merge_cells(f"A{row}:H{row}")
        row += 1
        col = 1
        for index, assignee in enumerate(self.assignees().values()):
            assignee_cell = sheet.cell(row, col)
            assignee_cell.value = assignee.name
            assignee_cell.font = Font(name="Helvetica", size=12)
            assignee_cell.border = border
            sheet.merge_cells(
                start_row=row,
                end_row=row,
                start_column=col,
                end_column=col + ((COLUMNS // sections) - 1),
            )
            col += 2
            if (index + 1) % sections == 0:
                col = 1
                row += 1
        ##################################
        if self.comments:
            row += 2
            comments_label = sheet.cell(row, 1)
            comments_label.value = "COMMENTS"
            comments_label.border = border
            comments_label.alignment = Alignment(horizontal="center")
            comments_label.fill = SUB_HEADER_BACKGROUND
            comments_label.font = SUB_HEADER_FONT
            sheet.merge_cells(f"A{row}:H{row}")
            row += 1
            comments_value = sheet.cell(row, 1)
            comments_value.value = self.comments
            comments_value.font = Font(name="Helvetica", size=12)
            comments_value.alignment = Alignment(horizontal="left", wrap_text=True)
            sheet.merge_cells(f"A{row}:H{row}")

            lines = comments_value.value.count("\n")
            row_height = (lines * 17) + 6
            sheet.row_dimensions[row].height = max(15.00, row_height)

            start_col, start_row, end_col, end_row = range_boundaries(f"A{row}:H{row}")
            for _row in sheet.iter_rows(
                min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col
            ):
                for cell in _row:
                    cell.border = border

            ##################################
            if self.close_out_comments:

                row += 2
                closeoutcomments_label = sheet.cell(row, 1)
                closeoutcomments_label.value = "CLOSE OUT COMMENTS"
                closeoutcomments_label.border = border
                closeoutcomments_label.alignment = Alignment(horizontal="center")
                closeoutcomments_label.fill = SUB_HEADER_BACKGROUND
                closeoutcomments_label.font = SUB_HEADER_FONT
                sheet.merge_cells(f"A{row}:H{row}")

                row += 1
                closeoutcomments_value = sheet.cell(row, 1)
                closeoutcomments_value.value = self.close_out_comments
                closeoutcomments_value.font = Font(name="Helvetica", size=12)
                closeoutcomments_value.alignment = Alignment(
                    horizontal="left", wrap_text=True
                )
                sheet.merge_cells(f"A{row}:H{row}")

                lines = closeoutcomments_value.value.count("\n")
                row_height = (lines * 17) + 6
                sheet.row_dimensions[row].height = max(15.00, row_height)

                start_col, start_row, end_col, end_row = range_boundaries(
                    f"A{row}:H{row}"
                )
                for _row in sheet.iter_rows(
                    min_row=start_row,
                    min_col=start_col,
                    max_row=end_row,
                    max_col=end_col,
                ):
                    for cell in _row:
                        cell.border = border

        ##################################
        if self.date_completed:
            row += 2
            datecompleted_label = sheet.cell(row, 1)
            datecompleted_label.value = "Date Completed"
            datecompleted_label.fill = SMALL_HEADER_BACKGROUND
            datecompleted_label.font = SMALL_HEADER_FONT
            datecompleted_label.alignment = Alignment(horizontal="right")
            datecompleted_label.border = border
            sheet.merge_cells(f"A{row}:B{row}")
            datecompleted_value = sheet.cell(row, 3)
            datecompleted_value.value = datetime.strftime(
                self.date_completed, "%d-%m-%Y"
            )
            datecompleted_value.font = VALUE_FONT
            datecompleted_value.border = border
            sheet.merge_cells(f"C{row}:H{row}")

        # Page Setup
        margin_size = 1 / 2.54
        sheet.page_margins.header = 0
        sheet.page_margins.left = margin_size
        sheet.page_margins.right = margin_size
        sheet.page_margins.top = margin_size
        sheet.page_margins.bottom = margin_size
        sheet.page_margins.footer = 0
        sheet.page_setup.fitToPage = True

        sheet.title = f"Work Order {self.id}"

        file_directory = None
        excel_file_name = None
        excel_file_path = None

        if not print_on_save:
            file_directory = select_directory()

            if not file_directory:
                return

            excel_file_name = f"{site.name} Work Order ({self.id}).xlsx"
            excel_file_path = rf"{file_directory}/{excel_file_name}"
        else:
            file_directory = os.path.expanduser("~/Documents")
            excel_file_name = f"TEMP{site.name} Work Order ({self.id}).xlsx"
            excel_file_path = rf"{file_directory}/{excel_file_name}"

        try:
            workbook.save(excel_file_path)
            SuccessNotification(
                "Save Successful",
                [f"Successfully saved work order to path: {excel_file_path}."],
            ).show()
        except Exception as e:
            ErrorNotification(
                "Save Failed",
                [
                    f"Save failed, check if file with same name ({excel_file_name}) is already open."
                ],
            ).show()
        finally:
            workbook.close()

        if print_on_save:
            excel = win32.Dispatch("Excel.Application")
            _print_workbook = excel.Workbooks.Open(excel_file_path)
            _print_worksheet = _print_workbook.ActiveSheet
            _print_worksheet.PrintOut()
            if excel.Workbooks.Count == 1:
                _print_workbook.Close(False)
                excel.Quit()
            else:
                _print_workbook.Close(False)
            os.remove(excel_file_path)
